import csv
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


MYSQL_EXE = r"D:\laragon\bin\mysql\mysql-8.4.3-winx64\bin\mysql.exe"
DB_NAME = "newtiffa_timesheet"
OUT_DIR = Path("exports")


def main():
    OUT_DIR.mkdir(exist_ok=True)
    out_file = OUT_DIR / f"presence_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    query = "SELECT * FROM presence ORDER BY flow_date DESC, id DESC;"
    cmd = [
        MYSQL_EXE,
        "-h",
        "127.0.0.1",
        "-P",
        "3307",
        "-u",
        "root",
        "--password=",
        "-D",
        DB_NAME,
        "--batch",
        "--raw",
        "-e",
        query,
    ]

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    reader = csv.reader(proc.stdout, delimiter="\t")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("presence")

    header = next(reader)
    ws.append(header)

    row_count = 0
    for row in reader:
        ws.append(row)
        row_count += 1

    stderr = proc.stderr.read()
    ret = proc.wait()
    if ret != 0:
        raise RuntimeError(stderr)

    meta = wb.create_sheet("metadata")
    meta.append(["database", DB_NAME])
    meta.append(["table", "presence"])
    meta.append(["exported_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["rows", row_count])

    wb.save(out_file)
    print(out_file.resolve())
    print(row_count)


if __name__ == "__main__":
    main()
