import csv
import subprocess
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"E:\VIBECODING\absen_tiffany")
MYSQL = Path(r"D:\laragon\bin\mysql\mysql-8.4.3-winx64\bin\mysql.exe")
SOURCES = [
    ("Sudirman", Path(r"E:\VIBECODING\absen_tiffany\sudirman.dat")),
    ("Gambir", Path(r"D:\Downloads\1_attlog (2).dat")),
]
OUTPUT = ROOT / "exports" / "laporan_absen_dat_gabungan.xlsx"


def mysql_tsv(query):
    result = subprocess.run(
        [str(MYSQL), "-uroot", "--batch", "--raw", "--skip-column-names", "newtiffa_timesheet", "-e", query],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return list(csv.reader(result.stdout.splitlines(), delimiter="\t"))


def get_employee_map():
    rows = mysql_tsv(
        """
        SELECT
            COALESCE(u.employee_code, ''),
            u.id,
            TRIM(CONCAT(COALESCE(u.first_name, ''), ' ', COALESCE(u.last_name, ''))),
            COALESCE(u.active, ''),
            COALESCE(b.id, ''),
            COALESCE(b.branch_code, ''),
            COALESCE(b.branch_name, ''),
            COALESCE(p.position_name, ''),
            COALESCE(s.subdivision_name, '')
        FROM users u
        LEFT JOIN position p ON p.id = u.position_id
        LEFT JOIN branch b ON b.id = p.branch_id
        LEFT JOIN subdivision s ON s.id = u.subdivision_id
        WHERE u.employee_code IS NOT NULL AND u.employee_code <> ''
        ORDER BY u.active DESC, b.id ASC, u.id ASC
        """
    )
    mapping = defaultdict(list)
    for row in rows:
        if len(row) < 9:
            continue
        employee = {
            "employee_code": row[0].strip(),
            "user_id": row[1],
            "employee_name": row[2],
            "active": row[3],
            "branch_id": row[4],
            "branch_code": row[5],
            "branch_name": row[6],
            "position_name": row[7],
            "subdivision_name": row[8],
        }
        mapping[employee["employee_code"]].append(employee)
    return mapping


def choose_employee(rows):
    if not rows:
        return None
    active = [row for row in rows if str(row.get("active")) == "1"]
    return active[0] if active else rows[0]


def parse_dat(label, path, employee_map):
    detail_rows = []
    grouped = defaultdict(lambda: {"times": [], "verify": defaultdict(int), "status": defaultdict(int), "workcode": defaultdict(int), "sources": set(), "employee": None})
    stats = {"source": label, "file": str(path), "lines": 0, "valid": 0, "invalid": 0}

    if not path.exists():
        stats["missing_file"] = 1
        return detail_rows, grouped, stats

    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            stats["lines"] += 1
            cols = line.split()
            if len(cols) < 6:
                stats["invalid"] += 1
                continue
            finger_id, date, time, verify, status, workcode = cols[:6]
            employee = choose_employee(employee_map.get(finger_id, []))
            branch = f"{employee['branch_code']} / {employee['branch_name']}" if employee and employee.get("branch_code") else "-"
            row = {
                "source": label,
                "finger_id": finger_id,
                "employee_name": employee["employee_name"] if employee else "TIDAK DITEMUKAN",
                "branch": branch,
                "position": employee.get("position_name") if employee else "-",
                "subdivision": employee.get("subdivision_name") if employee else "-",
                "active": employee.get("active") if employee else "-",
                "date": date,
                "time": time,
                "datetime": f"{date} {time}",
                "verify": verify,
                "status": status,
                "workcode": workcode,
            }
            detail_rows.append(row)
            group_key = (finger_id, date)
            grouped[group_key]["times"].append(time)
            grouped[group_key]["verify"][verify] += 1
            grouped[group_key]["status"][status] += 1
            grouped[group_key]["workcode"][workcode] += 1
            grouped[group_key]["sources"].add(label)
            grouped[group_key]["employee"] = employee
            stats["valid"] += 1
    return detail_rows, grouped, stats


def autosize(ws):
    for column_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def style_table(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def main():
    employee_map = get_employee_map()
    all_detail = []
    combined_grouped = defaultdict(lambda: {"times": [], "verify": defaultdict(int), "status": defaultdict(int), "workcode": defaultdict(int), "sources": set(), "employee": None})
    stats_rows = []

    for label, path in SOURCES:
        detail, grouped, stats = parse_dat(label, path, employee_map)
        all_detail.extend(detail)
        stats_rows.append(stats)
        for key, value in grouped.items():
            combined_grouped[key]["times"].extend(value["times"])
            combined_grouped[key]["sources"].update(value["sources"])
            combined_grouped[key]["employee"] = value.get("employee") or combined_grouped[key].get("employee")
            for bucket in ["verify", "status", "workcode"]:
                for item_key, count in value[bucket].items():
                    combined_grouped[key][bucket][item_key] += count

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Rekap Harian"
    ws_summary.append(["ID Finger", "Nama", "Cabang", "Posisi", "Sub Departemen", "Aktif", "Tanggal", "Scan Pertama", "Scan Terakhir", "Total Scan", "Verify", "Status", "WorkCode", "Sumber"])

    def sort_id(value):
        return int(value) if str(value).isdigit() else 999999

    for (finger_id, date), value in sorted(combined_grouped.items(), key=lambda item: (item[0][1], sort_id(item[0][0]), item[0][0])):
        times = sorted(value["times"])
        employee = value.get("employee")
        ws_summary.append([
            finger_id,
            employee["employee_name"] if employee else "TIDAK DITEMUKAN",
            f"{employee['branch_code']} / {employee['branch_name']}" if employee and employee.get("branch_code") else "-",
            employee.get("position_name") if employee else "-",
            employee.get("subdivision_name") if employee else "-",
            employee.get("active") if employee else "-",
            date,
            times[0] if times else "",
            times[-1] if times else "",
            len(times),
            ", ".join(f"{key}:{count}" for key, count in sorted(value["verify"].items())),
            ", ".join(f"{key}:{count}" for key, count in sorted(value["status"].items())),
            ", ".join(f"{key}:{count}" for key, count in sorted(value["workcode"].items())),
            ", ".join(sorted(value["sources"])),
        ])
    style_table(ws_summary)

    ws_detail = wb.create_sheet("Log Mentah")
    ws_detail.append(["Sumber", "ID Finger", "Nama", "Cabang", "Posisi", "Sub Departemen", "Aktif", "Tanggal", "Jam", "DateTime", "Verify", "Status", "WorkCode"])
    for row in sorted(all_detail, key=lambda item: (item["date"], item["time"], sort_id(item["finger_id"]), item["finger_id"])):
        ws_detail.append([row["source"], row["finger_id"], row["employee_name"], row["branch"], row["position"], row["subdivision"], row["active"], row["date"], row["time"], row["datetime"], row["verify"], row["status"], row["workcode"]])
    style_table(ws_detail)

    ws_stats = wb.create_sheet("Statistik")
    ws_stats.append(["Sumber", "File", "Total Baris", "Valid", "Invalid"])
    for row in stats_rows:
        ws_stats.append([row["source"], row["file"], row["lines"], row["valid"], row["invalid"]])
    ws_stats.append([])
    ws_stats.append(["Total Log Mentah", len(all_detail)])
    ws_stats.append(["Total Rekap ID+Tanggal", len(combined_grouped)])
    style_table(ws_stats)

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
