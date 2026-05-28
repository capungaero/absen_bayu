import csv
import subprocess
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"E:\VIBECODING\absen_tiffany")
MYSQL = Path(r"D:\laragon\bin\mysql\mysql-8.4.3-winx64\bin\mysql.exe")
SOURCES = [("Sudirman", Path(r"E:\VIBECODING\absen_tiffany\sudirman.dat")), ("Gambir", Path(r"D:\Downloads\1_attlog (2).dat"))]
OUTPUT = ROOT / "exports" / "laporan_absen_dat_rekap.xlsx"

def mysql_tsv(query):
    result = subprocess.run([str(MYSQL), "-uroot", "--batch", "--raw", "--skip-column-names", "newtiffa_timesheet", "-e", query], check=True, capture_output=True, text=True, encoding="utf-8", errors="replace")
    return list(csv.reader(result.stdout.splitlines(), delimiter="\t"))

def get_employee_map():
    rows = mysql_tsv("""
        SELECT COALESCE(u.employee_code,''), u.id, TRIM(CONCAT(COALESCE(u.first_name,''),' ',COALESCE(u.last_name,''))), COALESCE(u.active,''), COALESCE(b.id,''), COALESCE(b.branch_code,''), COALESCE(b.branch_name,''), COALESCE(p.position_name,''), COALESCE(s.subdivision_name,'')
        FROM users u
        LEFT JOIN position p ON p.id=u.position_id
        LEFT JOIN branch b ON b.id=p.branch_id
        LEFT JOIN subdivision s ON s.id=u.subdivision_id
        WHERE u.employee_code IS NOT NULL AND u.employee_code <> ''
        ORDER BY u.active DESC, b.id ASC, u.id ASC
    """)
    mapping = defaultdict(list)
    for row in rows:
        if len(row) < 9: continue
        item = {"employee_code": row[0].strip(), "user_id": row[1], "employee_name": row[2], "active": row[3], "branch_id": row[4], "branch_code": row[5], "branch_name": row[6], "position_name": row[7], "subdivision_name": row[8]}
        mapping[item["employee_code"]].append(item)
    return mapping

def choose_employee(rows):
    if not rows: return None
    active = [row for row in rows if str(row.get("active")) == "1"]
    return active[0] if active else rows[0]

def sort_id(value):
    return int(value) if str(value).isdigit() else 999999

def autosize(ws):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

def style_table(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)

employee_map = get_employee_map()
groups = defaultdict(lambda: {"times": [], "verify": defaultdict(int), "status": defaultdict(int), "workcode": defaultdict(int), "sources": set(), "employee": None})
stats = []
for label, path in SOURCES:
    line_count = valid = invalid = 0
    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            line = line.strip()
            if not line: continue
            line_count += 1
            cols = line.split()
            if len(cols) < 6:
                invalid += 1
                continue
            finger_id, date, time, verify, status, workcode = cols[:6]
            employee = choose_employee(employee_map.get(finger_id, []))
            key = (finger_id, date)
            groups[key]["times"].append(time)
            groups[key]["verify"][verify] += 1
            groups[key]["status"][status] += 1
            groups[key]["workcode"][workcode] += 1
            groups[key]["sources"].add(label)
            groups[key]["employee"] = employee or groups[key]["employee"]
            valid += 1
    stats.append((label, str(path), line_count, valid, invalid))

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = "Rekap Harian"
ws.append(["ID Finger", "Nama", "Cabang", "Posisi", "Sub Departemen", "Aktif", "Tanggal", "Scan Pertama", "Scan Terakhir", "Total Scan", "Verify", "Status", "WorkCode", "Sumber"])
for (finger_id, date), value in sorted(groups.items(), key=lambda item: (item[0][1], sort_id(item[0][0]), item[0][0])):
    times = sorted(value["times"])
    employee = value["employee"]
    ws.append([finger_id, employee["employee_name"] if employee else "TIDAK DITEMUKAN", f"{employee['branch_code']} / {employee['branch_name']}" if employee and employee.get("branch_code") else "-", employee.get("position_name") if employee else "-", employee.get("subdivision_name") if employee else "-", employee.get("active") if employee else "-", date, times[0], times[-1], len(times), ", ".join(f"{k}:{v}" for k, v in sorted(value["verify"].items())), ", ".join(f"{k}:{v}" for k, v in sorted(value["status"].items())), ", ".join(f"{k}:{v}" for k, v in sorted(value["workcode"].items())), ", ".join(sorted(value["sources"]))])
style_table(ws)

ws2 = wb.create_sheet("Statistik")
ws2.append(["Sumber", "File", "Total Baris", "Valid", "Invalid"])
for row in stats:
    ws2.append(list(row))
ws2.append([])
ws2.append(["Total Rekap ID+Tanggal", len(groups)])
style_table(ws2)
wb.save(OUTPUT)
print(OUTPUT)
